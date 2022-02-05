from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import sys

import pandas as pd
from datetime import date, datetime, timedelta
pd.set_option('display.max_columns', None)


def summarize_planner_export(sample_filepath):
    pre_process_result = pre_process(sample_filepath)
    post_process_result = post_processing(pre_process_result)
    format_final_result(post_process_result)
    return

def printCols(df):
    print(df.columns)
def pre_process(infile):

    # default columns
    # ['Task ID', 'Task Name', 'Bucket Name', 'Progress', 'Priority',
    #  'Assigned To', 'Created By', 'Created Date', 'Start Date', 'Due Date',
    #  'Late', 'Completed Date', 'Completed By', 'Description',
    #  'Completed Checklist Items', 'Checklist Items', 'Labels']
    # ['Task ID', 'Task Name', 'Bucket Name', 'Progress', 'Priority',
    #  'Assigned To', 'Created By', 'Created Date', 'Start Date', 'Due Date',
    #  'Late', 'Completed Date', 'Completed By', 'Description',
    #  'Completed Checklist Items', 'Checklist Items', 'Labels']
    # TODO: add functionality for Late items

    # check if file is open

    required_columns = ['Task Name', 'Priority', 'Assigned To', 'Due Date', 'Description']
    df = pd.read_excel(infile, usecols=required_columns)

    # remove items with no due date
    df = df[df['Due Date'].notna()]

    # convert due date from sting to datetime
    df['Due Date'] = df['Due Date'].astype(str)

    # filter to only the work items due today
    today = str(date.today().strftime("%m/%d/%Y"))
    due_today = df["Due Date"] == today
    pre_processed_df = df.loc[due_today]

    # tomorrow = date.today() + timedelta(days=1)
    # tomorrow = tomorrow.strftime("%m/%d/%Y")
    # due_tomorrow = df["Due Date"] == tomorrow


    print('pritning pre-provessed', )
    printCols(pre_processed_df)
    return pre_processed_df


def post_processing(dataframe):

    # Add Categories column
    post_processed_dataframe = dataframe

    # Remove Due Date Columns
    post_processed_dataframe.drop('Due Date', 1)

    # Clean up Assigned To column to only first names
    post_processed_dataframe['Assigned To'] = post_processed_dataframe['Assigned To'].str.replace(' [\w]*;', ', ', regex=True)
    post_processed_dataframe['Assigned To'] = post_processed_dataframe['Assigned To'].str.replace(' [\w]*$', '', regex=True)

    # TODO: remove populate Category column and drop Description column

    # Create custom sort order
    df_urgency_order = pd.DataFrame({
        'urgency': ['Urgent', 'Important', 'Medium', 'Low'],
    })
    sort_urgency = df_urgency_order.reset_index().set_index('urgency')

    # Create new column for sort order
    post_processed_dataframe['urgency_order'] = post_processed_dataframe['Priority'].map(sort_urgency['index'])

    # Sort by urgency_order
    post_processed_dataframe = post_processed_dataframe.sort_values('urgency_order')

    # then by Priority using custom sort 'Urgent', 'Important', 'Medium', 'Low'
    post_processed_dataframe = post_processed_dataframe.sort_values('Priority')

    return post_processed_dataframe


def format_final_result(dataframe):
    print(dataframe)
    today = str(date.today().strftime("%m_%d_%Y"))

    # export to excel file
    filename = f'Planner Daily Summary {today}.xlsx'

    ordered_columns = ['Task Name', 'Priority', 'Assigned To', 'Due Date', 'Description']
    dataframe.to_excel(filename, index=False, columns=ordered_columns)

    # bold top row
    wb = load_workbook(filename=filename)
    ws = wb['Sheet1']
    bold_font = Font(bold=True)

    # Enumerate the cells in the first row
    for cell in ws["1:1"]:
        cell.font = bold_font

    # update column widths
    column_widths = []
    for row in ws.iter_rows():
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if cell.value is not None:
                    if len(cell.value) > column_widths[i]:
                        column_widths[i] = len(cell.value)
            else:
                column_widths += [len(cell.value)]

    for i, column_width in enumerate(column_widths, 1):
        if i == 1:
            ws.column_dimensions[get_column_letter(i)].width = round(column_width)
        else:
            ws.column_dimensions[get_column_letter(i)].width = round(column_width * 1.2)

    wb.save(filename=filename)

    return

if __name__ == "__main__":
    SAMPLE_FILEPATH = sys.argv[1]
    summarize_planner_export(SAMPLE_FILEPATH)
